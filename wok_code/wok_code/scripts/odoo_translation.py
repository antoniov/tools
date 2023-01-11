#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Create map of Odoo modules
"""
import os
import sys
from time import time, sleep
import argparse
import re
import collections
from babel.messages import pofile
from openpyxl import load_workbook, Workbook

# from python_plus import unicodes

__version__ = "2.0.4"


MODULE_SEP = "\ufffa"

# (<en>, <it>, <module>, <result>)
TEST_DATA = [
    ("name", "nome", None, "nome"),
    ("Name", "Nome", None, "Nome"),
    ("Name.", "Name.", None, "Nome."),
    ("Name!!", "Name!!", None, "Nome!!"),
    ("First Name", "Nome di battesimo", None, "Nome di battesimo"),
    ("Last Name", "Cognome", None, "Cognome"),
    ("Name s.r.l.", None, None, "Nome s.r.l."),
    ("account", "contabilità", None, "contabilità"),
    ("tax", "IVA", None, "IVA"),
    ("UE", "EU", None, "EU"),
    ("Invoice", "Fattura", None, "Fattura"),
    ("invoice", "", None, "fattura"),
    ("<b>Invoice</b>", "<b>Invoice</b>", None, "<b>Fattura</b>"),
    ("<b>Invoice</b>", "", None, "<b>Fattura</b>"),
    ("<b>Invoice</b>", "<b>Fattura</b>", None, "<b>Fattura</b>"),
    ("Print <b>Invoice</b>!", "Stampa <b>Fattura</b>!", None, "Stampa <b>Fattura</b>!"),
    ("Order(s)", "Ordine/i", None, "Ordine/i"),
    ("Invoice n.%s", "", None, "Fattura n.%s"),
    ("Invoice n.%(number)s of %(date)s", "", None, "Fattura n.%(number)s of %(date)s"),
    ("Invoices", "", None, "Fatture"),
    ("Invoice(s)", "", None, "Fattura/e"),
    ("Invoice ", "", None, "Fattura "),
    (" Invoice", "", None, " Fattura"),
    (" Invoice ", "", None, " Fattura "),
    ("Invoices.", "", None, "Fatture."),
    ("line", "riga", None, "riga"),
    ("lines", "", None, "righe"),
    ("Invoice lines", "Righe fattura", None, "Righe fattura"),
    ("Sale", "Vendite", None, "Vendite"),
    ("Sale Invoice lines", "Righe fattura di vendita", None,
     "Righe fattura di vendita"),
    ("Credit", "Credito", None, "Credito"),
    ("Credit", "Avere", "l10n_it", "Avere"),
    ("account.tax", "*WRONG*", None, "account.tax"),
    ("Dear", "Gentile", None, "Gentile"),
    ("Dear ${name}", "", None, "Gentile ${name}"),
    ("Purchase", "", None, "Acquistare"),
    ("&gt; 100%%", "", None, "&gt; 100%%"),
    ("/usr/name/line", "", None, "/usr/name/line"),
    ("%s invoice lines", "righe %s fattura", None, "righe %s fattura"),
    ("# invoice lines", None, None, "Linee n.invoice"),
    ("Acceptance Account", "Conto RI.BA. all'incasso", None,
     "Conto RiBA all'incasso"),
    ("Ri.Ba. Bank", "Banca Ri.Ba", None, "Banca RiBA"),
]
msg_time = time()


def msg_burst(text):
    global msg_time
    t = time() - msg_time
    if t > 5:
        print(text, '\r')
        msg_time = time()


class OdooTranslation(object):
    """ """

    def __init__(self, opt_args):
        if sys.version_info[0] == 3:
            import translators as ts
        else:
            ts = None
        self.ts = ts
        self.opt_args = opt_args
        self.dict = {}
        if (
            opt_args.target_path
            and opt_args.target_path.endswith(".po")
            and os.path.basename(os.path.dirname(opt_args.target_path)) == "i18n"
        ):
            self.opt_args.target_path = os.path.dirname(
                os.path.dirname(opt_args.target_path))
        if not opt_args.file_xlsx:
            root = self.get_home_devel()
            if opt_args.dbg_template:
                dict_name = os.path.join(root, 'pypi', 'tools', 'odoo_default_tnl.xlsx')
            else:
                dict_name = os.path.join(root, 'venv', 'bin', 'odoo_default_tnl.xlsx')
            if os.path.isfile(dict_name):
                self.opt_args.file_xlsx = dict_name

        # Type classification
        # Name, Translatable, Grouped, Regex, Nolast
        self.types_decl = [
            ("odoo_model",
             False,
             False,
             (
                 r"^([a-z0-9]{2,}[\._][a-z0-9]{2,}([\._][a-z0-9]{2,})?"
                 r"|[0-9]+|/[-\w._]{2,}(/[-\w._]{2,})+)"
             ),
             False),
            ("word", True, True, (
                r"([a-zA-Z]{1,2}[./]([a-zA-Z]{1,2}[./])*([a-zA-Z]{1,2})*"
                r"|(\w|# )(-?\w|\w| # |`)*)"),
             False),
            ("punct", True, False, r"[.,:;!?()]+", True),
            ("space", False, True, r"\s+", True),
            ("tag", False, False, (
                r"(</?\w+[^/>]*/?>|%[a-zA-Z]|%\(\w+\)[a-zA-Z]"
                r"|/w+(/w+)*|\$\{[^}]+\}|&\w+;)"
            ), False),
            ("space2", False, True, r"['’\"«»&]", True),
        ]
        for item in self.types_decl:
            if item[0] == "word":
                self.re_word = item[3]
            elif item[0] == "space":
                self.re_space = item[3]
            elif item[0] == "space2":
                self.re_space2 = item[3]
            elif item[0] == "tag":
                self.re_tag = item[3]
        self.build_alias_dict()

    def get_home_devel(self):
        root = os.environ.get('HOME_DEVEL')
        if not root or not os.path.isdir(root):
            if os.path.isdir(os.path.expanduser('~/odoo/devel')):
                root = os.path.expanduser('~/odoo/devel')
            elif os.path.isdir(os.path.expanduser('~/devel')):
                root = os.path.expanduser('~/devel')
            else:
                root = os.environ.get('HOME')
        return root

    def ismodule(self, path):
        if os.path.isdir(path):
            if (
                os.path.isfile(os.path.join(path, "__manifest__.py"))
                or os.path.isfile(os.path.join(path, "__openerp__.py"))
            ) and os.path.isfile(os.path.join(path, "__init__.py")):
                return True
        return False

    def isplural(selfself, term):
        return (term.endswith("s")
                and not term.endswith("%s")
                and not term.endswith(")s")) or term.endswith("(s)")

    def get_filenames(self, filename=None):
        filename = filename or self.opt_args.file_xlsx
        tmp_file = bak_file = None
        if filename:
            tmp_file = "%s.tmp" % filename
            bak_file = "%s.bak" % filename
        return filename, tmp_file, bak_file

    def save_n_bak_fn(self, filename, tmp_file, bak_file):
        if os.path.isfile(bak_file):
            os.unlink(bak_file)
        os.rename(filename, bak_file)
        os.rename(tmp_file, filename)

    def get_hash_key(self, key, ignore_case, module=None):
        kk = key.strip()
        if ignore_case:
            kk = key if key == key.upper() else key.lower()
        kk2 = ""
        if module:
            kk2 = module + MODULE_SEP + kk
        return kk, kk2

    def adjust_case(self, orig, tnxl):
        if tnxl:
            if orig == orig.upper():
                tnxl = tnxl.upper()
            elif len(tnxl) > 1:
                if orig[0].isupper() and tnxl[0].islower():
                    tnxl = tnxl[0].upper() + tnxl[1:]
                elif orig[0].islower() and tnxl[0].isupper() and tnxl != tnxl.upper():
                    tnxl = tnxl[0].lower() + tnxl[1:]
        return tnxl

    def set_plural(self, orig, term):
        def plural_term(term, suffix, sep):
            if sep:
                term += (sep + suffix)
            else:
                term = term[: -1] + suffix
            return term

        sep = "/" if orig.endswith("(s)") else ""
        if term.endswith("ca") or term.endswith("ga"):
            term = plural_term(term, "he", sep)
        elif term.endswith("cia") or term.endswith("gia"):
            if term[-4] in ("a", "e", "i", "o", "u"):
                term = plural_term(term, "ie", sep)
            else:
                term = plural_term(term, "e", sep)
        elif term.endswith("a"):
            term = plural_term(term, "e", sep)
        elif term.endswith("e") or term.endswith("o"):
            term = plural_term(term, "i", sep)
        return term

    def set_plural_if(self, hash_key, orig, tnxl, adjust_case=None):
        if self.isplural(orig):
            hkey = hash_key[: -1]
            if hkey in self.dict:
                tnxl = self.adjust_case(
                    orig,
                    self.set_plural(orig, self.dict[hkey][1]))
            elif adjust_case:
                tnxl = self.adjust_case(orig, tnxl)
        elif adjust_case:
            tnxl = self.adjust_case(orig, tnxl)
        return tnxl

    def get_term(self, hash_key, orig, tnxl, adjust_case=None):
        tnxl = tnxl or orig
        if orig:
            if hash_key in self.dict:
                tnxl = self.adjust_case(orig, self.dict[hash_key][1])
            else:
                tnxl = self.set_plural_if(hash_key, orig, tnxl, adjust_case=adjust_case)
        return tnxl

    def store_1_item(
        self, hash_key, msg_orig, msg_tnxl, override=None, module=None, is_tag=None
    ):
        if len(msg_orig) <= 1 and not is_tag:
            return msg_orig
        if collections.Counter(msg_orig)['%'] != collections.Counter(msg_tnxl)['%']:
            print("*** Warning: different macro: " + msg_orig + " / " + msg_tnxl)
        if not is_tag:
            for item in self.tags:
                tnxls = self.dict[item]
                ltoken = "%s " % tnxls[0]
                rtoken = "%s " % tnxls[1]
                if msg_tnxl.startswith(ltoken):
                    msg_tnxl = msg_tnxl.replace(ltoken, rtoken, 1)
                ltoken = " %s " % tnxls[0]
                rtoken = " %s " % tnxls[1]
                msg_tnxl = msg_tnxl.replace(ltoken, rtoken)
                ltoken = " %s" % tnxls[0]
                rtoken = " %s" % tnxls[1]
                if msg_tnxl.endswith(ltoken):
                    msg_tnxl = msg_tnxl[0: -len(ltoken)] + rtoken
        hash_key, hashkey_mod = self.get_hash_key(hash_key, False, module=module)
        if hashkey_mod and (
            hashkey_mod not in self.dict
            or override
            or (msg_tnxl and self.dict[hashkey_mod][0] == self.dict[hashkey_mod][1])
        ):
            self.dict[hashkey_mod] = (msg_orig, msg_tnxl)
        elif hash_key and (
            hash_key not in self.dict
            or override
            or is_tag
            or (msg_tnxl and self.dict[hash_key][0] == self.dict[hash_key][1])
        ):
            self.dict[hash_key] = (msg_orig, msg_tnxl)
            if is_tag and hash_key not in self.tags:
                self.tags.append(hash_key)
        return self.get_term(hash_key, msg_orig, msg_tnxl)

    def split_items(self, message):
        def append_group(tokens, hash_keys, groups, hash_groups):
            if re.search(self.re_space, groups[-1]):
                tokens.append(groups[: -1])
                tokens.append(groups[-1])
                hash_keys.append(hash_groups[: -1])
                hash_keys.append("")
            else:
                tokens.append(groups)
                hash_keys.append(hash_groups)
            groups = []
            hash_groups = []
            return tokens, hash_keys, groups, hash_groups

        tokens = []
        groups = []
        hash_keys = []
        hash_groups = []
        ix = 0
        while ix < len(message):
            for tok_type in self.types_decl:
                match = re.match(tok_type[3], message[ix:])
                if match:
                    token = message[ix: match.end() + ix]
                    if token.startswith("# "):
                        token = "N."
                        hash_key = self.get_hash_key(token, tok_type[0])[0]
                        grouped = tok_type[2]
                        ix += 2
                    else:
                        hash_key = self.get_hash_key(token, tok_type[0])[0]
                        grouped = tok_type[2]
                        ix += match.end()
                    break
            if not match:
                ii = len(message) - ix
                grouped = False
                for tok_type in self.types_decl:
                    match = re.search(tok_type[3], message[ix:])
                    if match:
                        ii = min(ii, match.start())
                        break
                tok_type = None
                token = message[ix: ii + ix]
                hash_key = self.get_hash_key(token, False)[0]
                ix += ii
            if (
                (re.search(self.re_space, token) or re.search(self.re_space2, token))
                and not groups
            ):
                grouped = False
            if grouped:
                groups.append(token)
                hash_groups.append(hash_key)
            else:
                if groups:
                    tokens, hash_keys, groups, hash_groups = append_group(
                        tokens, hash_keys, groups, hash_groups)
                tokens.append(token)
                if tok_type and tok_type[4] and ix >= len(message):
                    hash_keys.append("")
                else:
                    hash_keys.append(hash_key)
        if groups:
            tokens, hash_keys, groups, hash_groups = append_group(
                tokens, hash_keys, groups, hash_groups)
        return tokens, hash_keys

    def do_dict_item(
        self, msg_orig, msg_tnxl, action=None, override=None, module=None, is_tag=None,
    ):
        action = action or ("build_dict" if override else "translate")
        for tok_type in self.types_decl:
            if not tok_type[1]:
                if re.fullmatch(tok_type[3], msg_orig):
                    return msg_orig
        if not msg_tnxl:
            msg_tnxl = msg_orig
        msg_orig = msg_orig.replace("’", "'")
        msg_tnxl = msg_tnxl.replace("’", "'")
        tokens_orig, hashes_orig = self.split_items(msg_orig)
        tokens_tnxl, hashes_tnxl = self.split_items(msg_tnxl)
        try_to_translate = True if len(hashes_orig) == len(hashes_tnxl) else False
        ix = 0
        while try_to_translate and ix < len(hashes_orig):
            if (
                (
                    isinstance(hashes_orig[ix], (list, tuple))
                    and not isinstance(hashes_tnxl[ix], (list, tuple))
                )
                or
                (
                    not isinstance(hashes_orig[ix], (list, tuple))
                    and isinstance(hashes_tnxl[ix], (list, tuple))
                )
                or
                (
                    len(hashes_orig[ix]) != len(hashes_tnxl[ix])
                )
                or
                (
                    all([
                        (len(hashes_orig[ix][x]) <= 2
                         or hashes_orig[ix][x] != hashes_tnxl[ix][x])
                        for x in range(len(hashes_orig[ix]))])
                )
            ):
                try_to_translate = False
            ix += 1
        fullterm_orig = fullterm_tnxl = fulltermhk_orig = fulltermhk_tnxl = ""
        hash_key = ""
        tok_orig = tokens_orig.pop(0) if tokens_orig else ""
        hash_orig = hashes_orig.pop(0) if hashes_orig else ""
        tok_tnxl = tokens_tnxl.pop(0) if tokens_tnxl else ""
        fullterm_2_store = False
        adjust_case = True
        while tok_orig or tok_tnxl:
            if isinstance(tok_orig, (list, tuple)):
                term_orig = "".join(tok_orig)
                hkey = "".join(hash_orig)
                if module:
                    hkey = self.get_hash_key(hkey, True, module=module)[1]
                fullterm_orig += term_orig
                hash_key += hkey
                fulltermhk_orig = fullterm_orig
                if isinstance(tok_tnxl, (list, tuple)):
                    term_tnxl = self.get_hash_key(
                        "".join(tok_tnxl), not adjust_case, module=module)[0]
                    if (
                        (action == "build_dict" or not try_to_translate)
                        and hkey not in self.dict
                        and self.get_hash_key(term_orig,
                                              not adjust_case,
                                              module=module)[0] != term_tnxl
                    ):
                        self.store_1_item(hkey, term_orig, term_tnxl)
                        fullterm_2_store = True
                    else:
                        x = self.get_term(
                            hkey, term_orig, term_tnxl, adjust_case=adjust_case)
                        if x.lower() != term_tnxl.lower():
                            term_tnxl = x
                            fullterm_2_store = True
                        else:
                            x = ""
                            ctr = 0
                            for ii, term in enumerate(tok_orig):
                                if (
                                    len(term) > 3
                                    and re.match(self.re_word, term)
                                    and hash_orig[ii] in self.dict
                                ):
                                    x += self.get_term(hash_orig[ii],
                                                       term,
                                                       term,
                                                       adjust_case=adjust_case)
                                    adjust_case = False
                                    ctr += 1
                                else:
                                    x += term
                            if ctr == 1:
                                term_tnxl = x
                                fullterm_2_store = True
                        adjust_case = False
                    fullterm_tnxl += term_tnxl
                    fulltermhk_tnxl = fullterm_tnxl
                    if isinstance(tok_tnxl, (list, tuple)):
                        tok_tnxl = tokens_tnxl.pop(0) if tokens_tnxl else ""
                tok_orig = tokens_orig.pop(0) if tokens_orig else ""
                hash_orig = hashes_orig.pop(0) if hashes_orig else ""
            elif tok_orig:
                if tok_orig.endswith(".") or tok_orig.endswith("-"):
                    adjust_case = True
                fullterm_orig += tok_orig
                hash_key += hash_orig
                while tok_tnxl and not isinstance(tok_tnxl, (list, tuple)):
                    fullterm_tnxl += tok_tnxl
                    tok_tnxl = tokens_tnxl.pop(0) if tokens_tnxl else ""
                if hash_orig:
                    fulltermhk_orig = fullterm_orig
                    fulltermhk_tnxl = fullterm_tnxl
                tok_orig = tokens_orig.pop(0) if tokens_orig else ""
                hash_orig = hashes_orig.pop(0) if hashes_orig else ""
            else:
                if isinstance(tok_tnxl, (list, tuple)):
                    fullterm_tnxl += "".join(tok_tnxl)
                else:
                    fullterm_tnxl += tok_tnxl
                fulltermhk_tnxl = fullterm_tnxl
                tok_tnxl = tokens_tnxl.pop(0) if tokens_tnxl else ""
                fullterm_2_store = True

        for tok_type in self.types_decl:
            if not tok_type[1]:
                if re.fullmatch(tok_type[3], fullterm_orig):
                    return fullterm_orig
        if action == "build_dict":
            if is_tag or (msg_orig != msg_tnxl and msg_tnxl):
                return self.store_1_item(
                    hash_key, msg_orig, msg_tnxl,
                    override=override, module=module, is_tag=is_tag)
            elif self.isplural(fullterm_tnxl):
                fulltermhk_tnxl = self.set_plural(fullterm_orig, fullterm_tnxl[: -3])
                fullterm_2_store = True
            if fullterm_orig.lower() == fullterm_tnxl.lower() or not fullterm_tnxl:
                if self.ts and not re.search(self.re_tag, fullterm_orig):
                    try:
                        # Use Google translator
                        fullterm_tnxl = self.adjust_case(
                            fullterm_orig,
                            self.ts.google(fullterm_orig,
                                           from_language='en',
                                           to_language=self.opt_args.lang[:2],
                                           timeout=5))
                        sleep(0.2)
                    except BaseException:
                        pass
                return self.store_1_item(
                    hash_key, fullterm_orig, fullterm_tnxl, override=override)
            elif fullterm_orig and fullterm_2_store:
                return self.store_1_item(
                    hash_key, fulltermhk_orig, fulltermhk_tnxl,
                    override=override, module=module)
        elif (
            action == "translate"
            and fullterm_tnxl == fulltermhk_tnxl
            and hash_key in self.dict
        ):
            return self.get_term(hash_key, fullterm_orig, fullterm_tnxl)
        elif fullterm_2_store or action == "translate":
            return self.set_plural_if(hash_key, fullterm_orig, fullterm_tnxl)
        return fullterm_orig

    def translate_item(self, msg_orig, msg_tnxl, module=None):
        return self.do_dict_item(msg_orig, msg_tnxl, action="translate", module=module)

    def translate_pofile(self, po_fn):
        def add_po_line(potext, line):
            potext += line
            potext += "\n"
            return potext

        def try_to_sync(left_lines, right_lines, left_ix, right_ix):
            if left_ix < len(left_lines) and right_ix < len(right_lines):
                saved_ix = right_ix
                ctr = 6
                while ctr and left_lines[left_ix] != right_lines[right_ix]:
                    right_ix += 1
                    if right_ix < len(right_lines):
                        ctr -= 1
                    else:
                        ctr = 0
                        right_ix -= 1
                if left_lines[left_ix] == right_lines[right_ix]:
                    return left_ix, right_ix, True

                right_ix = saved_ix
                saved_ix = left_ix
                ctr = 6
                while ctr and left_lines[left_ix] != right_lines[right_ix]:
                    left_ix += 1
                    if left_ix < len(left_lines):
                        ctr -= 1
                    else:
                        ctr = 0
                        left_ix -= 1
                if left_lines[left_ix] == right_lines[right_ix]:
                    return left_ix, right_ix, True
                left_ix = saved_ix
            return left_ix, right_ix, False

        def find_diff(left_lines, right_lines, left_ix, right_ix):
            while left_ix < len(left_lines) and right_ix < len(right_lines):
                if (
                    left_lines[left_ix] == right_lines[right_ix]
                    or (
                    re.match(r"^\".*\"$", left_lines[left_ix])
                    and re.match(r"^\".*\"$", right_lines[right_ix])
                )):
                    left_ix += 1
                    right_ix += 1
                    continue
                elif re.match(r"^\".*\"$", left_lines[left_ix]):
                    left_ix += 1
                    continue
                elif re.match(r"^\".*\"$", right_lines[right_ix]):
                    right_ix += 1
                    continue

                if left_lines[left_ix].startswith("#:"):
                    break

                left_ix, right_ix, equals = try_to_sync(
                    left_lines, right_lines, left_ix, right_ix)
                if equals:
                    continue

                if (
                    left_lines[left_ix].startswith("msgid")
                    and right_lines[right_ix].startswith("msgid")
                ):
                    right_ix += 1
                    left_ix += 1
                    continue
                if (
                    left_lines[left_ix].startswith("msgstr")
                    and right_lines[right_ix].startswith("msgstr")
                ):
                    right_ix += 1
                    left_ix += 1
                    continue
                break
            return left_ix, right_ix

        if os.path.isfile(po_fn):
            filename, tmp_file, bak_file = self.get_filenames(filename=po_fn)
            if self.opt_args.verbose:
                print("Writing %s" % filename)

            module = self.opt_args.module_name
            try:
                catalog = pofile.read_po(open(po_fn, "r"))
            except BaseException as e:
                print("Error %s reading po file %s" % (e, po_fn))
                return
            for message in catalog:
                if not message.id:
                    continue
                message.string = self.translate_item(
                    message.id, message.string, module=module)

            pofile.write_po(open(tmp_file, 'wb'),
                            catalog,
                            include_previous=True)
            with open(filename, "r") as fd:
                left_lines = fd.read().split("\n")
            with open(tmp_file, "r") as fd:
                right_lines = fd.read().split("\n")
            updated = False
            left_ix = right_ix = 0
            while left_ix < len(left_lines):
                left_ix, right_ix = find_diff(
                    left_lines, right_lines, left_ix, right_ix)
                if left_ix >= len(left_lines):
                    break
                if (right_ix < len(right_lines)
                    and left_lines[left_ix].startswith("#:")
                    and (not right_lines[right_ix].startswith("#:")
                         or left_lines[left_ix].split(":")[1]
                         !=
                         right_lines[right_ix].split(":")[1])
                ):
                    updated = True
                    right_lines.insert(right_ix, left_lines[left_ix])
                else:
                    saved_ix = left_ix
                    left_ix, right_ix, equals = try_to_sync(
                        left_lines, right_lines, left_ix, right_ix)
                    if left_ix != saved_ix:
                        left_ix = saved_ix
                        updated = True
                        right_lines.insert(right_ix, left_lines[left_ix])
                left_ix += 1
                right_ix += 1

            LAST_TNL_NAME = 'Antonio M. Vigliotti'
            LAST_TNL_MAIL = 'antoniomaria.vigliotti@gmail.com'
            LAST_TEAM_NAME = 'Zeroincombenze'
            LAST_TEAM_URL = 'https://www.zeroincombenze.it'
            LANGUAGE = r'"Language: it_IT\n"'
            PLURALS = r'"Plural-Forms: nplurals=2; plural=(n != 1);\n"'
            lang_it_lineno = False
            plurals_lineno = False
            for lineno, line in enumerate(right_lines):
                if re.match(r"^#\s+\*",  line):
                    right_lines[lineno] = r"# * %s" % module
                elif line.startswith('#. module:'):
                    right_lines[lineno] = r"#. module: %s" % module
                elif (
                    self.opt_args.odoo_branch
                    and line.startswith('"Project-Id-Version:')
                ):
                    right_lines[lineno] = (
                        r'"Project-Id-Version: Odoo (%s)\n"'
                        % self.opt_args.odoo_branch
                    )
                elif line.startswith('"Last-Translator:'):
                    right_lines[lineno] = (
                        r'"Last-Translator: %s <%s>\n"'
                        % (LAST_TNL_NAME, LAST_TNL_MAIL)
                    )
                elif line.startswith('"Language-Team:'):
                    right_lines[lineno] = (
                        r'"Language-Team: %s (%s)\n"'
                        % (LAST_TEAM_NAME, LAST_TEAM_URL)
                    )
                    if not lang_it_lineno:
                        lang_it_lineno = lineno
                    if not plurals_lineno:
                        plurals_lineno = lineno
                elif line.startswith('"Language:'):
                    right_lines[lineno] = LANGUAGE
                    lang_it_lineno = -1
                elif line.startswith('"Plural-Forms:'):
                    right_lines[lineno] = PLURALS
                    plurals_lineno = -1
                elif (
                    line.startswith('#')
                    and lineno > 0
                    and right_lines[lineno] == right_lines[lineno - 1]
                ):
                    del right_lines[lineno]
                    lineno -= 1
            if plurals_lineno and plurals_lineno >= 0:
                right_lines.insert(plurals_lineno, PLURALS)
                updated = True
            if lang_it_lineno and lang_it_lineno >= 0:
                right_lines.insert(lang_it_lineno, LANGUAGE)
                updated = True

            if updated:
                with open(tmp_file, "w") as fd:
                    fd.write("\n".join(right_lines))

            self.save_n_bak_fn(filename, tmp_file, bak_file)

    def load_terms_from_pofile(self, po_fn, override=None):
        if os.path.isfile(po_fn):
            if self.opt_args.verbose:
                print("Loading %s" % po_fn)
            try:
                catalog = pofile.read_po(open(po_fn, "r"))
            except BaseException as e:
                print("Error %s reading po file %s" % (e, po_fn))
                return
            for message in catalog:
                if not message.id:
                    continue
                self.do_dict_item(message.id,
                                  message.string,
                                  action="build_dict",
                                  override=override)

    def load_terms_from_xlsx(self, dict_fn):
        if os.path.isfile(dict_fn):
            if self.opt_args.verbose:
                print("Loading terms from %s" % dict_fn)
            wb = load_workbook(dict_fn)
            sheet = wb.active
            colnames = []
            for ncol in sheet.columns:
                colnames.append(ncol[0].value)
            hdr = True
            for nrow in sheet.rows:
                if hdr:
                    hdr = False
                    continue
                row = {}
                for ncol, cell in enumerate(nrow):
                    row[colnames[ncol]] = (
                        cell.value.replace("\\n", "\n") if cell.value else cell.value
                    )
                if not row["msgid"] or not row["msgstr"]:
                    continue
                msg_burst('%s ...' % row["msgid"])
                if "hashkey" in row and row["hashkey"]:
                    if row["hashkey"] not in self.dict:
                        self.dict[row["hashkey"]] = (row["msgid"], row["msgstr"])
                else:
                    self.do_dict_item(row["msgid"],
                                      row["msgstr"],
                                      action="build_dict",
                                      module=row["module"])

    def load_terms_for_test(self):
        for items in TEST_DATA:
            self.do_dict_item(items[0], items[1], action="build_dict", module=items[2])

    def do_work_on_path(self, root, base, action=None):
        action = action or "translate"
        path = os.path.join(root, base) if base else root
        if self.ismodule(path):
            i18n_path = os.path.join(path, "i18n")
            po_fn = os.path.join(i18n_path, "%s.po" % self.opt_args.lang)
            if not os.path.isfile(po_fn):
                po_fn = os.path.join(
                    i18n_path, "%s.po" % self.opt_args.lang.split("_")[0]
                )
            if not os.path.isfile(po_fn):
                print("Module %s without translation" % os.path.basename(path))
                return
            if action == "build_dict":
                self.load_terms_from_pofile(po_fn)
            elif action == "translate":
                self.translate_pofile(po_fn)
            else:
                raise

    def build_alias_dict(self):
        self.tags = []
        for (msg_orig, msg_tnxl, is_tag) in (
            ("I", "io", True),
            ("a", "un", False),
            ("iva", "IVA", True),
            ("sepa", "SEPA", True),
            ("UE", "EU", True),
            ("Ri.Ba.", "RiBA", True),
            ("Ri.Ba", "RiBA", True),
            ("RI.BA.", "RiBA", True),
            ("DDT", "DdT", True),
            ("ddt", "DdT", True),
        ):
            self.do_dict_item(msg_orig,
                              msg_tnxl,
                              action="build_dict",
                              is_tag=True)

    def build_dict(self):
        if self.opt_args.file_xlsx:
            root = self.get_home_devel()
            if root:
                dict_name = os.path.join(
                    root, 'pypi', 'tools', 'odoo_template_tnl.xlsx')
                if os.path.isfile(dict_name):
                    self.load_terms_from_xlsx(dict_name)
            self.load_terms_from_xlsx(self.opt_args.file_xlsx)
        if not self.opt_args.module_name:
            target_path = os.path.abspath(self.opt_args.target_path)
            self.do_work_on_path(target_path, None)
            for root, dirs, files in os.walk(target_path,
                                             topdown=True,
                                             followlinks=False):
                dirs[:] = [
                    d
                    for d in dirs
                    if d not in (".git", "__to_remove", "doc", "setup", ".idea")
                ]
                for base in dirs:
                    self.do_work_on_path(root, base, action="build_dict")

    def translate_module(self):
        module = self.opt_args.module_name
        target_path = os.path.abspath(self.opt_args.target_path)
        if module == "OCB" and os.path.isfile(os.path.join(target_path, "odoo-bin")):
            self.do_work_on_path(target_path, None, action="translate")
        elif module == os.path.basename(target_path):
            self.do_work_on_path(target_path, None, action="translate")
        else:
            for root, dirs, files in os.walk(target_path,
                                             topdown=True,
                                             followlinks=False):
                dirs[:] = [
                    d
                    for d in dirs
                    if d not in (".git", "__to_remove", "doc", "setup", ".idea")
                ]
                for base in dirs:
                    if module == base:
                        self.do_work_on_path(root, base, action="translate")
                        break

    def list_dict(self):
        for hash_key, terms in self.dict.items():
            print("[%s]\n'%s'='%s'" % (hash_key, terms[0], terms[1]))

    def write_xlsx(self):
        if os.path.isfile(self.opt_args.file_xlsx):
            filename, tmp_file, bak_file = self.get_filenames()
            if self.opt_args.verbose:
                print("Writing %s" % filename)
            wb = Workbook()
            sheet = wb.active
            sheet.title = "odoo_default_tnxl"
            sheet.cell(row=1, column=1, value="module")
            sheet.cell(row=1, column=2, value="msgid")
            sheet.cell(row=1, column=3, value="msgstr")
            sheet.cell(row=1, column=4, value="hashkey")
            row = 1
            for hash_key, terms in self.dict.items():
                row += 1
                if MODULE_SEP in hash_key:
                    module = hash_key.split(MODULE_SEP)[0]
                else:
                    module = ""
                sheet.cell(row=row, column=1, value=module)
                sheet.cell(row=row, column=2, value=terms[0])
                sheet.cell(row=row, column=3, value=terms[1])
                sheet.cell(row=row, column=4, value=hash_key)
            wb.save(tmp_file)
            self.save_n_bak_fn(filename, tmp_file, bak_file)


def main(cli_args=None):
    cli_args = cli_args or sys.argv[1:]
    parser = argparse.ArgumentParser(
        description="Create translation file", epilog="© 2022-2023 by SHS-AV s.r.l."
    )
    parser.add_argument(
        '-B', '--debug-template', action='store_true', dest='dbg_template'
    )
    parser.add_argument(
        "-b",
        "--odoo-branch",
        dest="odoo_branch",
        default="12.0",
        help="Default Odoo version",
    )
    parser.add_argument("-c", "--config", help="Odoo configuration file")
    parser.add_argument(
        "-G",
        "--git-orgs",
        help="Git organizations, comma separated - " "May be: oca librerp or zero",
    )
    parser.add_argument("-l", "--lang", default="it_IT", help="Language")
    parser.add_argument("-m", "--module-name")
    parser.add_argument("-n", "--dry-run", action="store_true")
    parser.add_argument("-p", "--target-path", help="Local directory")
    parser.add_argument("-T", "--test", action="store_true")
    parser.add_argument("-v", "--verbose", action="count", default=0)
    parser.add_argument("-V", "--version", action="version", version=__version__)
    parser.add_argument("-W", "--rewrite-xlsx",
                        action="store_true",
                        help="Rewrite xlsx file")
    parser.add_argument("-x", "--file-xlsx", help="Default dictionary")

    odoo_tnxl = OdooTranslation(parser.parse_args(cli_args))
    if odoo_tnxl.opt_args.test:
        odoo_tnxl.load_terms_for_test()
    elif odoo_tnxl.opt_args.file_xlsx or odoo_tnxl.opt_args.target_path:
        odoo_tnxl.build_dict()
    else:
        print("Invalid parameters: please set xlsx file or Odoo path")
        return 1
    if odoo_tnxl.opt_args.module_name:
        odoo_tnxl.translate_module()
    if odoo_tnxl.opt_args.rewrite_xlsx:
        odoo_tnxl.write_xlsx()
    if odoo_tnxl.opt_args.test:
        print("")
        print("")
        print("Test starting ...")
        ctr = ctr_err = 0
        for items in TEST_DATA:
            res = odoo_tnxl.translate_item(items[0], items[1], module=items[2])
            print("[%s]->[%s]" % (items[0], items[3]))
            if items[3] != res:
                print("    **** TEST FAILED!!!! [" + items[3] + "]!=[" + res + "]")
                ctr_err += 1
            ctr += 1
        print("")
        for items in odoo_tnxl.dict.values():
            if items[0] != items[0].strip():
                print("    **** TEST FAILED!!!! [" + items[0] + "] with trailing space")
                ctr_err += 1
            ctr += 1
        print("")
        print("%d test executed with %d error detected." % (ctr, ctr_err))
    return 0


if __name__ == "__main__":
    exit(main())
